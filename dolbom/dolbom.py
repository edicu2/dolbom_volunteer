# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'main.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!
import webbrowser

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QStyleFactory, QTableWidgetItem
from collections import Counter
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import calendar
import math


# 높이만 변경하고 나머지 같은 돌보미는 같은 색으로 표시하는것만 해주기
import os



class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(958, 495)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.table = QtWidgets.QTableWidget(self.centralwidget)
        self.table.setGeometry(QtCore.QRect(10, 50, 938, 291))
        self.table.setRowCount(50000)
        self.table.setColumnCount(7)
        self.table.setObjectName("table")
        self.table.setMinimumSize(450, 400)
        self.table.resizeColumnsToContents()
        self.table.setSizeAdjustPolicy(
            QtWidgets.QAbstractScrollArea.AdjustToContents)

        for i in range(0, 7):
            item = QtWidgets.QTableWidgetItem()
            item.setBackground(QtGui.QColor(88, 88, 88))
            brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
            brush.setStyle(QtCore.Qt.SolidPattern)
            item.setForeground(brush)
            self.table.setHorizontalHeaderItem(i, item)

        for i in range(0,1000):
            item = QtWidgets.QTableWidgetItem()
            item.setTextAlignment( Qt.AlignVCenter | Qt.AlignCenter)
            self.table.setItem(i, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
            self.table.setItem(i, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
            self.table.setItem(i, 2, item)




        self.btn_change_excel = QtWidgets.QPushButton(self.centralwidget)
        self.btn_change_excel.setIcon(QtGui.QIcon('excel_icon.png'))
        self.btn_change_excel.setIconSize(QtCore.QSize(100,31))
        self.btn_change_excel.setGeometry(QtCore.QRect(720, 10, 100, 31))
        self.btn_change_excel.setObjectName("btn_change_excel")
        self.btn_change_excel.clicked.connect(self.change_excel)

        self.btn_paste2 = QtWidgets.QPushButton(self.centralwidget)
        self.btn_paste2.setGeometry(QtCore.QRect(130, 10, 111, 31))
        self.btn_paste2.setObjectName("btn_paste_2")
        self.btn_paste2.clicked.connect(self.paste_table)


        self.btn_change_html = QtWidgets.QPushButton(self.centralwidget)
        self.btn_change_html.setGeometry(QtCore.QRect(830, 10, 100, 31))
        self.btn_change_html.setIcon(QtGui.QIcon('internet_icon.png'))
        self.btn_change_html.setIconSize(QtCore.QSize(100, 31))
        self.btn_change_html.setObjectName("btn_change_html")

        self.btn_paste = QtWidgets.QPushButton(self.centralwidget)
        self.btn_paste.setGeometry(QtCore.QRect(10, 10, 111, 31))
        self.btn_paste.setObjectName("btn_paste")
        self.btn_paste.clicked.connect(self.paste_excel)


        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 958, 21))
        self.menubar.setObjectName("menubar")
        self.menuprogram = QtWidgets.QMenu(self.menubar)
        self.menuprogram.setObjectName("menuprogram")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.action_3 = QtWidgets.QAction(MainWindow)
        self.action_3.setObjectName("action_3")
        self.menuprogram.addAction(self.action_3)
        self.menubar.addAction(self.menuprogram.menuAction())

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "돌보미 급여 서비스"))
        MainWindow.setWindowIcon(QtGui.QIcon('logo.png'))
        item = self.table.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "돌보미 명"))
        item = self.table.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "활동 시작"))
        item = self.table.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "활동 종료"))
        item = self.table.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "이용자"))
        item = self.table.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "아동명"))
        item = self.table.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "요일"))
        item = self.table.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "이용시간"))
        __sortingEnabled = self.table.isSortingEnabled()


        #table Colum width 지정하기
        self.table.setColumnWidth(0, 110)
        self.table.setColumnWidth(1,160)
        self.table.setColumnWidth(2,160)
        self.table.setColumnWidth(3,110)
        self.table.setColumnWidth(4,110)
        self.table.setColumnWidth(5,120)
        self.table.setColumnWidth(6,120)
        self.table.setSortingEnabled(True)
        self.table.setSortingEnabled(__sortingEnabled)

        self.btn_change_excel.setText(_translate("MainWindow", ""))
        self.btn_change_html.setText(_translate("MainWindow", ""))
        self.btn_paste.setText(_translate("MainWindow", "데이터 삽입/추가"))
        self.btn_paste2.setText(_translate("MainWindow", "데이터 적용"))
        self.menuprogram.setTitle(_translate("MainWindow", "도움말"))
        self.action_3.setText(_translate("MainWindow", "설명서"))

    def paste_excel(self):

        webbrowser.open("form.xlsx")

    def paste_table(self):
        excel_document = openpyxl.load_workbook('form.xlsx')
        sheet = excel_document["A샘플-주간"]
        #sheet = excel_document["B샘플-야간"]
        row_len = sheet.max_row              # 길이가 인식 안됨
        self.table.clearContents()
        for idx in range(4,row_len+1):

            print(idx)
            teacher = sheet.cell(row=idx, column=3).value
            item = QTableWidgetItem(str(teacher))
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
            self.table.setItem(idx-4, 0, item)

            start_ac = sheet.cell(row=idx, column=14).value
            item = QTableWidgetItem(str(start_ac))
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
            self.table.setItem(idx - 4, 1, item)

            end_ac = sheet.cell(row=idx, column=15).value
            item = QTableWidgetItem(str(end_ac))
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
            self.table.setItem(idx - 4, 2, item)

            client =  sheet.cell(row=idx, column=7).value
            item = QTableWidgetItem(str(client))
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
            self.table.setItem(idx - 4, 3,item)

            child = sheet.cell(row=idx, column=11).value
            item = QTableWidgetItem(str(child))
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
            self.table.setItem(idx - 4, 4, item)

            day_of_week = sheet.cell(row=idx, column=16).value
            item = QTableWidgetItem(str(day_of_week))
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignCenter)
            self.table.setItem(idx - 4, 5, item)

            total_time = sheet.cell(row=idx, column=17).value
            item = QTableWidgetItem(str(total_time)+" 시간")
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
            self.table.setItem(idx - 4, 6, item)
        excel_document.close()

    def change_excel(self):
        # 원본이 있는 excel 파일
        excel = openpyxl.load_workbook('form.xlsx')
        sheet = excel["A샘플-주간"]
        #sheet = excel["B샘플-야간"]


        #데이터 변환할 excel 파일
        excel2 = openpyxl.load_workbook('output.xlsx')
        sheet2 = excel2["달력형식"]

        # 달력 날짜랑 달 만들기 위해서 사용 #
        searchYear = sheet.cell(row=4, column=14).value[0:4]
        searchMonth = sheet.cell(row=4, column=14).value[5:7]
        searchName = sheet.cell(row=4, column=3).value
        startDayOfWeek = calendar.monthrange(int(searchYear), int(searchMonth))[0]
        totalDay = calendar.monthrange(int(searchYear), int(searchMonth))[1]
        totalWeekend = round((startDayOfWeek + totalDay) / 7 + 0.5)

        # 선생님 수 파악하기
        teacher = [sheet["C4"].value]
        row_len = sheet.max_row
        for idx in range(4, row_len + 1):
            if not sheet["C"+str(idx)].value in teacher:
                teacher.append(sheet["C"+str(idx)].value)
        print(teacher)
        ## 달력 excel에 날짜 기입
        # 3 ,5 ,7 ,9, 11 ,13 여기에 날짜

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        t_count = 0
        for teach_n in teacher:
            t_count += 1
            date = 0
            sheet2["C"+str((t_count-1)*19+1)] = searchYear + " / " + searchMonth + "     " + teach_n[0:3]
            sheet2["C" + str((t_count - 1) * 19 + 1)].font = Font(size= 30 , color ="FF8282", bold= True)
            sheet2["C" + str((t_count - 1) * 19 + 2)] = "월"
            sheet2["C" + str((t_count - 1) * 19 + 2)].font = Font(size= 16 , bold= True)
            sheet2["D" + str((t_count - 1) * 19 + 2)] = "화"
            sheet2["D" + str((t_count - 1) * 19 + 2)].font = Font(size=16, bold=True)
            sheet2["E" + str((t_count - 1) * 19 + 2)] = "수"
            sheet2["E" + str((t_count - 1) * 19 + 2)].font = Font(size=16, bold=True)
            sheet2["F" + str((t_count - 1) * 19 + 2)] = "목"
            sheet2["F" + str((t_count - 1) * 19 + 2)].font = Font(size=16, bold=True)
            sheet2["G" + str((t_count - 1) * 19 + 2)] = "금"
            sheet2["G" + str((t_count - 1) * 19 + 2)].font = Font(size=16, bold=True)
            sheet2["H" + str((t_count - 1) * 19 + 2)] = "토"
            sheet2["H" + str((t_count - 1) * 19 + 2)].font = Font(size=16, bold=True)
            sheet2["I" + str((t_count - 1) * 19 + 2)] = "일"
            sheet2["I" + str((t_count - 1) * 19 + 2)].font = Font(size=16, color="FF3636", bold=True)
            sheet2["J" + str((t_count - 1) * 19 + 2)] = "합계"
            sheet2["J" + str((t_count - 1) * 19 + 2)].font = Font(size=16, bold=True)

            input = True
            for i in range(3,14,2):
                for j in range(3,10):
                    if date == totalDay:
                        date = 0
                        input = False
                    sheet2.row_dimensions[i + (t_count - 1) * 19 + 1].height = 100
                    sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19)].border = thin_border
                    sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)].border = thin_border
                    if ((i == 3 and startDayOfWeek <= j - 3) or (i > 3)):
                        date += 1
                        sheet2[str(chr(j+64))+str(i+(t_count-1)*19)] = str(date)
                        sheet2[str(chr(j+64)) + str(i + (t_count - 1) * 19)].font = Font(size=14)
                        sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 +1) ].alignment = Alignment(wrapText=True, vertical="top")
                        if j == 9 :
                            sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19)].font = Font(size=14, color="FF3636")
                        if input == False :
                            sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19)].font = Font(size=14, color="BDBDBD")
                sheet2["J"+ str(i + (t_count - 1) * 19)].border = thin_border
                sheet2["J" + str(i + (t_count - 1) * 19)].alignment = Alignment(wrapText=True, vertical="top")
                sheet2["J" + str(i + (t_count - 1) * 19 + 1)].border = thin_border
                sheet2["J" + str(i + (t_count - 1) * 19 + 1)].alignment = Alignment(wrapText=True, vertical="top")
        # 기입 끝

        # 데이터전부 list로 날짜에 맞게 집어넣기
        date_time = {}
        row_len = sheet.max_row
        teacher = ""
        for i in range(4, row_len + 1):
            teacher = self.teacher(sheet, i)

            # 처음 시작할 때 dict 생성
            if not teacher in date_time :
                date_time[teacher] = {}
            if (self.time_to_num(sheet, "N", i) < 0): #  이 말은 앞 뒤 날짜가 다르다는 말
                start_time = 24 + self.time_to_num(sheet, "N", i)
                end_time = self.time_to_num(sheet, "O", i)
                start_date = self.day(sheet, "N", i)
                end_date = self.day(sheet, "O", i)
                if (self.month(sheet,"O",i) != int(searchMonth)):
                    end_date = 32
                if not start_date in date_time[teacher]: date_time[teacher][start_date] = []
                if not end_date in date_time[teacher]: date_time[teacher][end_date] = []
                for i in self.myrange(float(start_time), 24.25, 0.25):
                    date_time[teacher][start_date].append(i)

                for i in self.myrange(float(0),end_time+0.25,0.25):
                    date_time[teacher][end_date].append(i)

            else :
                start_time = self.time_to_num(sheet, "N", i)
                end_time = self.time_to_num(sheet, "O", i)
                start_date = self.day(sheet, "N", i)
                if not start_date in date_time[teacher]: date_time[teacher][start_date] = []
                for i in self.myrange(float(start_time), end_time+0.25, 0.25):
                    date_time[teacher][start_date].append(i)



        #########################################################################
        #  아이가 중복 시간을 계산
        #  날짜별로 구분
        #########################################################################
        date_duple = {}
        duple = 0
        duple_num = 0
        for teacher in date_time.keys():  # 선생님 이름
            date_duple[teacher] = {}
            for i in date_time[teacher].keys():  # 선생님 일한 날짜
                length = 0
                for key in Counter(date_time[teacher][i]): length += 1
                # length 길이 구하기
                len_count = 0
                duple_data = 0
                why = "no"
                for key in Counter(date_time[teacher][i]):
                    len_count += 1
                    if not i in date_duple[teacher]:
                        date_duple[teacher][i] = {}
                        duple = Counter(date_time[teacher][i])[key]
                        duple_num = 0
                        duple_data = key

                    if not Counter(date_time[teacher][i])[key] in date_duple[teacher][i]:
                        date_duple[teacher][i][Counter(date_time[teacher][i])[key]] = 0

                    if duple != Counter(date_time[teacher][i])[key] or duple_data != key:
                        if why == "dupleBottom" :
                            if (duple_data != key):
                                why = "jump"
                                date_duple[teacher][i][duple] += (duple_num) * 0.25
                            else :
                                if(duple < Counter(date_time[teacher][i])[key]):
                                    why = "dupleTop"
                                    date_duple[teacher][i][duple] += (duple_num+1) * 0.25
                                else:
                                    why = "dupleBottom"
                                    date_duple[teacher][i][duple] += (duple_num) * 0.25
                        else :
                            if (duple_data != key):
                                why = "jump"
                                date_duple[teacher][i][duple] += (duple_num-1) * 0.25
                            else :
                                if(duple < Counter(date_time[teacher][i])[key]):
                                    why = "dupleTop"
                                    date_duple[teacher][i][duple] += (duple_num) * 0.25
                                else:
                                    why = "dupleBottom"
                                    date_duple[teacher][i][duple] += (duple_num -1) * 0.25


                        duple_data = key
                        duple = Counter(date_time[teacher][i])[key]
                        #print("key: 중복횟수 : ", key, Counter(date_time[i])[key],date_duple[i][duple])
                        duple_num = 0


                    if (len_count == length):
                        if why == "dupleBottom" :
                            date_duple[teacher][i][duple] += (duple_num +1) * 0.25
                        else :
                            date_duple[teacher][i][duple] += (duple_num) * 0.25
                        #print("key: 중복횟수 : ", key, Counter(date_time[i])[key],date_duple[i][duple])

                    duple_data += 0.25
                    duple_num += 1



        #########################################################################
        # 야간 중복 찾기
        # date 아이 중복이 있는지 시간 계산 함
        #########################################################################
        night_duple = {}
        night_totalTime = {}
        duple = 0
        duple_num = 0
        for teacher in date_time.keys():  # 선생님 이름
            night_duple[teacher] = {}
            night_totalTime[teacher] = {}
            for i in date_time[teacher].keys():
                # length 길이 구하기
                length = 0
                for key in Counter(date_time[teacher][i]): length += 1
                len_count = 0
                duple_data = 0
                why = "no"
                for key in Counter(date_time[teacher][i]):
                    len_count += 1
                    if not i in night_duple[teacher]:
                        night_totalTime[teacher][i] = 0
                        night_duple[teacher][i] = {}
                        duple = Counter(date_time[teacher][i])[key]
                        duple_num = 0
                        duple_data = key

                    if not Counter(date_time[teacher][i])[key] in night_duple[teacher][i]:
                        night_duple[teacher][i][Counter(date_time[teacher][i])[key]] = 0

                    if duple != Counter(date_time[teacher][i])[key] or duple_data != key:
                        if duple_num == 0 :
                            print(" ")
                        elif why == "dupleBottom":
                            if (duple_data != key):
                                why = "jump"
                                night_duple[teacher][i][duple] += (duple_num) * 0.25
                            else:
                                if (duple < Counter(date_time[teacher][i])[key]):
                                    why = "dupleTop"
                                    night_duple[teacher][i][duple] += (duple_num + 1) * 0.25
                                else:
                                    why = "dupleBottom"
                                    night_duple[teacher][i][duple] += (duple_num) * 0.25
                        else:
                            if (duple_data != key):
                                why = "jump"
                                night_duple[teacher][i][duple] += (duple_num - 1) * 0.25
                            else:
                                if (duple < Counter(date_time[teacher][i])[key]):
                                    why = "dupleTop"
                                    night_duple[teacher][i][duple] += (duple_num) * 0.25
                                else:
                                    why = "dupleBottom"
                                    night_duple[teacher][i][duple] += (duple_num - 1) * 0.25
                        night_totalTime[teacher][i] += night_duple[teacher][i][duple]
                        duple_data = key
                        duple = Counter(date_time[teacher][i])[key]
                        #print("key: 중복횟수 : ", key, Counter(date_time[i])[key], night_duple[i][duple])
                        duple_num = 0

                    if (len_count == length):
                        if why == "dupleBottom":
                            night_duple[teacher][i][duple] += (duple_num + 1) * 0.25
                        else:
                            night_duple[teacher][i][duple] += (duple_num) * 0.25
                        night_totalTime[teacher][i] = night_duple[teacher][i][duple]
                        #print("key: 중복횟수 : ", key, Counter(date_time[i])[key], night_duple[i][duple])
                    duple_data += 0.25
                    if ((key>=22.0 and key<=24.0) or (key>=0 and key<=6.0 )):
                        duple_num += 1



        #########################################################################
        # 8시간 이상 초과 근무인지 찾기
        # 8시간 이상이고 야간이면 2배
        # - 몇시간 초과인지 중복당 추가해주기
        # 시간이 0~ 22시 사이의 시간에서 찾기
        #########################################################################
        eight_duple = {}
        duple_duple = {}
        duple_totalTime = {}
        eight_totalTime = {}
        eight_bool = False
        duple = 0
        duple_num = 0
        for teacher in date_time.keys():  # 선생님 이름
            duple_duple[teacher] = {}
            eight_duple[teacher] = {}
            duple_totalTime[teacher] = {}
            eight_totalTime[teacher] = {}
            for i in date_time[teacher].keys():
                length = 0
                for key in Counter(date_time[teacher][i]): length += 1
                # length 길이 구하기
                len_count = 0
                duple_data = 0
                duple2_data = 0
                eight_bool = False
                eight_time = 0
                why = "no"
                for key in Counter(date_time[teacher][i]):
                    len_count += 1
                    if not i in eight_duple[teacher]:
                        duple_duple[teacher][i] = {}
                        duple_totalTime[teacher][i] = 0
                        eight_totalTime[teacher][i] = 0
                        eight_duple[teacher][i] = {}
                        duple = Counter(date_time[teacher][i])[key]
                        duple_num = 0
                        duple_data = key
                        duple2_data = key
                    if not Counter(date_time[teacher][i])[key] in eight_duple[teacher][i]:
                        eight_duple[teacher][i][Counter(date_time[teacher][i])[key]] = 0
                        duple_duple[teacher][i][Counter(date_time[teacher][i])[key]] = 0

                    if duple != Counter(date_time[teacher][i])[key] or duple_data != key:
                        if why == "dupleBottom":
                            if (duple_data != key):
                                why = "jump"
                                eight_duple[teacher][i][duple] += (duple_num) * 0.25
                            else:
                                if (duple < Counter(date_time[teacher][i])[key]):
                                    why = "dupleTop"
                                    eight_duple[teacher][i][duple] += (duple_num + 1) * 0.25
                                else:
                                    why = "dupleBottom"
                                    eight_duple[teacher][i][duple] += (duple_num) * 0.25
                        else:
                            if (duple_data != key):
                                why = "jump"
                                eight_duple[teacher][i][duple] += (duple_num - 1) * 0.25
                            else:
                                if (duple < Counter(date_time[teacher][i])[key]):
                                    why = "dupleTop"
                                    eight_duple[teacher][i][duple] += (duple_num) * 0.25
                                else:
                                    why = "dupleBottom"
                                    eight_duple[teacher][i][duple] += (duple_num - 1) * 0.25
                        eight_time += eight_duple[teacher][i][duple]


                        if(eight_time>= 8):
                            if(eight_bool == False):
                                eight_duple[teacher][i][duple] = eight_time - 8
                                eight_bool = True
                        else : eight_duple[teacher][i][duple] = 0

                        if(eight_bool == True):
                            if duple_data > 22.0:
                                print("경우 1")
                                if duple2_data>= 22.0:
                                    eight_duple[teacher][i][duple] -= (duple_data-0.25 - duple2_data)
                                    duple_duple[teacher][i][duple] += (duple_data-0.25 - duple2_data)
                                else:
                                    print("경우 2")
                                    eight_duple[teacher][i][duple] -= (duple_data-0.25 - 22.0)
                                    duple_duple[teacher][i][duple] += (duple_data-0.25 - 22.0)
                        eight_totalTime[teacher][i] += eight_duple[teacher][i][duple]
                        duple_totalTime[teacher][i] += duple_duple[teacher][i][duple]

                        duple_data = key
                        duple2_data = key
                        duple = Counter(date_time[teacher][i])[key]
                        #print("key: 중복횟수 : ", key, Counter(date_time[i])[key], eight_duple[i][duple])
                        duple_num = 0

                    if (len_count == length):
                        #print(key)
                        if why == "dupleBottom":
                            eight_duple[teacher][i][duple] += (duple_num + 1) * 0.25
                        else:
                            eight_duple[teacher][i][duple] += (duple_num) * 0.25
                        #print("key: 중복횟수 : ", key, Counter(date_time[i])[key], eight_duple[i][duple])
                        if (eight_time >= 8):
                            if (eight_bool == False):
                                eight_duple[teacher][i][duple] = eight_time - 8
                                eight_bool = True
                        else:
                            eight_duple[teacher][i][duple] = 0
                        if (eight_bool == True ):
                            print(duple_data, duple2_data, "경우 3")
                            if duple_data > 22.0:
                                if duple2_data>= 22.0:
                                    eight_duple[teacher][i][duple] -= (duple_data - duple2_data)
                                    duple_duple[teacher][i][duple] += (duple_data - duple2_data)
                                else:
                                    eight_duple[teacher][i][duple] -= (duple_data - 22.0)
                                    duple_duple[teacher][i][duple] += (duple_data - 22.0)
                        eight_totalTime[teacher][i] += eight_duple[teacher][i][duple]
                        duple_totalTime[teacher][i] += duple_duple[teacher][i][duple]
                        break
                    duple_data += 0.25
                    duple_num += 1




        # 달력 중복시간 기입
        # 3 ,5 ,7 ,9, 11 ,13 여기에 날짜
        t_count = 0
        for teacher in date_duple:
            t_count += 1
            date = 0
            input = True
            for i in range(3, 14, 2):
                for j in range(3, 10):
                    if date == totalDay:
                        date = 0
                        input = False
                    if ((i == 3 and startDayOfWeek <= j - 3) or (i > 3 and input) or (date == 0 and input == False)):
                        date += 1
                        if (input == False):
                            if 32 in date_time[teacher]:
                                if date in date_duple[teacher]:
                                    for du in range(1,4):
                                        duple_text = du+"명 중복: "
                                        if du in date_duple[teacher][date]: duple_text = duple_text + str(date_duple[teacher][date][du])+" H /"
                                        else : duple_text = duple_text + " - H /"
                                        if du in eight_duple[teacher][date]: duple_text = duple_text + str(eight_duple[teacher][date][du])+" H /"
                                        else : duple_text = duple_text + " - H /"
                                        if du in night_duple[teacher][date]: duple_text = duple_text + str(night_duple[teacher][date][du])+" H /"
                                        else : duple_text = duple_text + " - H  /"
                                        if du in duple_duple[teacher][date]: duple_text = duple_text + str(duple_duple[teacher][date][du])+" H /"
                                        else : duple_text = duple_text + " - H \n"
                                        if sheet2[str(chr(j + 64)) + str(i+ (t_count - 1) * 19 + 1)].value in [None,'None'] : sheet2[str(chr(j + 64)) + str(i+ (t_count - 1) * 19 + 1)] = duple_text
                                        else : sheet2[str(chr(j + 64)) + str(i+ (t_count - 1) * 19 + 1)] = str(sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)].value) + "\n" + duple_text
                                    print(date , sheet2[str(chr(j + 64)) + str(i+ (t_count - 1) * 19 + 1)].value)
                                break

                        if date in date_duple[teacher]:
                            print(date_duple[teacher][date].keys(), " key 값")
                            for du in range(1, 4):
                                duple_text = str(du) + "명 중복: "
                                if du in date_duple[teacher][date]:
                                    duple_text = duple_text + str(date_duple[teacher][date][du]) + " H /"
                                else:
                                    duple_text = duple_text + " 0 H /"
                                if du in eight_duple[teacher][date]:
                                    duple_text = duple_text + str(eight_duple[teacher][date][du]) + " H /"
                                else:
                                    duple_text = duple_text + " 0 H /"
                                if du in night_duple[teacher][date]:
                                    duple_text = duple_text + str(night_duple[teacher][date][du]) + " H /"
                                else:
                                    duple_text = duple_text + " 0 H /"
                                if du in duple_duple[teacher][date]:
                                    duple_text = duple_text + str(duple_duple[teacher][date][du]) + " H /"
                                else:
                                    duple_text = duple_text + " 0 H \n"
                                if sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)].value in [None, 'None']:
                                    sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)] = duple_text
                                else:
                                    sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)] = str(sheet2[str(
                                        chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)].value) + "\n" + duple_text
                            print(date, sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)].value)

        ## 달력 시간 기입
        date_totalTime = {}
        week_totalTime = {}
        totalTime = {}
        for teacher in date_duple.keys():
            date_totalTime[teacher] = {}
            week_totalTime[teacher] = {}
            totalTime[teacher] = 0
            for i in date_duple[teacher].keys():
                for j in date_duple[teacher][i].items():
                    week = self.week(int(i), startDayOfWeek)
                    #print(week)
                    #print(i)
                    if week in week_totalTime[teacher]: week_totalTime[teacher][week] += j[1]
                    else : week_totalTime[teacher][week] = j[1]

                    if i in date_totalTime[teacher]: date_totalTime[teacher][i] += j[1]
                    else : date_totalTime[teacher][i] = j[1]
                    totalTime[teacher] +=j[1]
                #print("날짜 ",i,date_totalTime[teacher])
                #print("주 ", week, week_totalTime[teacher])
            #print(totalTime[teacher])


        # date , week 총 금액 출력
        t_count = 0
        week_money = {}    # 배열만 들기
        for teacher in date_duple.keys():   # 선생님 구분 해서 week 넣기
            t_count += 1
            date = 0                        # 선생님이 변경 될 때마다 date = 0 초기화
            input = True                    # input true 로 해서 마지막날이 지날 때 False로 변경
            for i in range(3, 14, 2):               # 달력에 넣기 위해  해당 양식 사용
                if not teacher in week_money :      # week_money에 teacher이 없다면 넣기
                    week_money[teacher] = {}
                for j in range(3, 10):              # 한 주 입력
                    total_money = 0                 # total_money = 0으로 변경
                    if date == totalDay:            # 전체날짜랑 같아지면 date=  0
                        date = 0                    # input을 False로 변경
                        input = False
                    if ((i == 3 and startDayOfWeek <= j - 3) or (i > 3 and input) or (date == 0 and input == False)):  # 달력에 맞게 계산한 것 넣기
                        date += 1
                        if (input == False):
                            if 32 in date_totalTime[teacher]:   # 32일은 다음달로 넘어간 첫번째 날로 입력
                                duple_text = "총: "
                                duple_text = duple_text + str(date_totalTime[teacher][date]) + " H /"
                                duple_text = duple_text + str(eight_totalTime[teacher][date]) + " H /"
                                duple_text = duple_text + str(night_totalTime[teacher][date]) + " H /"
                                duple_text = duple_text + str(duple_totalTime[teacher][date]) + " H \n"


                                if sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)].value in [None,'None']:
                                    sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)] = duple_text
                                else:
                                    sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)] =  duple_text + str(sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)].value)

                        if(date in date_totalTime[teacher]):    # 최대 날짜를 넘지 않은 경우

                            duple_text = "총: "
                            duple_text = duple_text + str(date_totalTime[teacher][date]) + " H /"
                            duple_text = duple_text + str(eight_totalTime[teacher][date]) + " H /"
                            duple_text = duple_text + str(night_totalTime[teacher][date]) + " H /"
                            duple_text = duple_text + str(duple_totalTime[teacher][date]) + " H \n"
                            if sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)].value in [None, 'None']:
                                sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)] = duple_text
                            else:
                                sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)] = duple_text + str(
                                    sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)].value)

                                # 달력에 맞춰서 데이터 입력 하루 총 시간을 입력
                            #sheet2[str(chr(j + 64)) + str(i + 1)] = "총  " + str(date_totalTime[teacher][date]) + " H" + "\n" + str(sheet2[str(chr(j + 64)) + str(i + 1)].value)
                        if date in date_duple[teacher]:            # 중복하는 아동별로 계산 시작
                            for k in date_duple[teacher][date].keys():  # ex> 1명 2명 3명 4명  key값을 가져옴
                                #print(date,k)
                                #print( "계싼", (date_duple[date][k] - night_duple[date][k]+ eight_duple[date][k])*8400 + (night_duple[date][k] + eight_duple[date][k])*10080)
                                # 계산 시작
                                if k == 1:
                                    total_money += (date_duple[teacher][date][k] - night_duple[teacher][date][k]- eight_duple[teacher][date][k] - duple_duple[teacher][date][k])*8400+ ((night_duple[teacher][date][k] + eight_duple[teacher][date][k])*8400*1.5) + duple_duple[teacher][date][k]*8400*2
                                    #print(total_money, "1 일 때")
                                if k == 2:
                                    total_money += (date_duple[teacher][date][k] - night_duple[teacher][date][k]- eight_duple[teacher][date][k] - duple_duple[teacher][date][k])*8400*1.5+ ((night_duple[teacher][date][k] + eight_duple[teacher][date][k])*8400*1.5*1.5) + duple_duple[teacher][date][k]*8400*2*1.5
                                    #print(total_money, "2 일 때")
                                if k == 3:
                                    total_money += (date_duple[teacher][date][k] - night_duple[teacher][date][k]- eight_duple[teacher][date][k] - duple_duple[teacher][date][k])*8400* 2.0\
                                                   + (night_duple[teacher][date][k] + eight_duple[teacher][date][k]) * 8400 *2.0*1.5 -(night_duple[teacher][date][k] + eight_duple[teacher][date][k]) *20*2.0*1.5
                                    #print(total_money, "3 일 때")
                            #print(total_money)         # 계산이 끝나고나면 총 total금액이 나온다.
                            week = self.week(date, startDayOfWeek)
                            if not week in week_money[teacher]: week_money[teacher][week] = 0
                            week_money[teacher][week] += total_money   # 선생님 week_money에 넣는다
                            sheet2[str(chr(j + 64 )) + str(i + (t_count - 1) * 19 + 1)] = "총 금액  " + str(int(total_money)) + " 원" + "\n" + str(sheet2[str(chr(j + 64)) + str(i + (t_count - 1) * 19 + 1)].value)

        t_count = 0
        for teacher in date_duple.keys():
            t_count += 1
            for w in week_totalTime[teacher].keys():
                #print(w)
                sheet2["J" + str((w*2)+2+(t_count - 1) * 19 )].value = "총 : " + str(week_totalTime[teacher][w]) + "H"
                sheet2["J" + str((w*2)+2+(t_count - 1) * 19 )].value ="총 금액 "+str(int(week_money[teacher][w])) +" 원\n"+ str(sheet2["J"+str((w*2) + 2+(t_count - 1) * 19 )].value)



        saveName = self.file_save()[0]
        excel2.save(saveName + ".xlsx")

        if saveName :
            excel3 = openpyxl.load_workbook(saveName + ".xlsx")
            sheet3 = excel3["표형식"]

            #sheet3 표형식에 출력
            row = 1
            for teacher in date_duple.keys():
                totalMoney = 0
                totalTime = 0
                row += 1
                sheet3["B" + str(row)].value = teacher.split(",")[0]+")"
                for week in range(1,7) :
                    if week in week_totalTime[teacher]:
                        sheet3[chr(66+week) + str(row)] = str(week_totalTime[teacher][week])
                        sheet3[chr(66+week) + str(row)] = str(sheet3[chr(66+week) + str(row)].value) + " // "+ "{:,}".format(int(week_money[teacher][week]))
                        totalTime+= week_totalTime[teacher][week]
                        totalMoney += week_money[teacher][week]
                    else :
                        sheet3[chr(66+week) + str(row)] = " - "
                if 32 in date_totalTime[teacher]:
                    sheet3["J" + str(row)].value = date_totalTime[teacher][32]
                else :
                    sheet3["J" + str(row)].value = "  - "
                sheet3["I" + str(row)].value = str(totalTime) + " / " + "{:,}".format(int(totalMoney))

            excel3.save(saveName + ".xlsx")

        '''
    


        for i in week_totalTime.items():
            print(i)

        for i in date_totalTime.items():
            print(i)
        '''
        '''
        for i in date_totalTime.items():
            print("date_total",i)
            print("----------------------")

        for i in week_totalTime.items():
            print("week_total",i)
            print("----------------------")

        for i in week_totalTime.keys():
            sheet2["J" + str((i * 2) + 2)] = "총 : " + str(week_totalTime[i]) + "H"
        '''

    def file_save(self):
        fileName = QtWidgets.QFileDialog.getSaveFileName(None, '저장할 위치 선택', "C:\\")
        return fileName

    def myrange(self,start, end, step):
        while start < end:
            yield start
            start += step
    def teacher(self,sheet,num):
        return sheet["C" + str(num)].value
    def client(self,sheet,num):
        return sheet["G"+str(num)].value

    def month(self,sheet,col,num):
        return int(sheet[col+str(num)].value[5:7])

    def day(self,sheet,col,num):
        return int(sheet[col+str(num)].value[8:10])

    def week(self ,date, startDayOfWeek):
        return int(math.ceil((date + startDayOfWeek) / 7))

    def time_to_num(self,sheet,char,num):
        x = 0
        if (char == "N"):
            if (sheet["N"+str(num)].value[8:10] != sheet["O" + str(num)].value[8:10]):
                x = int(sheet["N" + str(num)].value[11:13]) - 24  # 다음표의 이용자가 같으면 일단 start 지점과 end지점을 표시
            else :
                x = int(sheet["N" + str(num)].value[11:13])
            if (sheet["N" + str(num)].value[14:15] != "0"):  # 데이터를 넣어준다.
                x += 0.5
            return x
        else :
            x = int(sheet["O" + str(num)].value[11:13])
            if (sheet["O" + str(num)].value[14:15] != "0"):  # 데이터를 넣어준다.
                x += 0.5
            return x







if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle(QStyleFactory.create('Fusion'))
    app.setWindowIcon(QtGui.QIcon('logo.png'))
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

